from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db import models
from django.utils import timezone
from django.utils.timezone import localtime
from datetime import timedelta
from .models import Producto, Categoria, Historial
from .forms import ProductoForm, UsuarioForm
from openpyxl.styles import Font, PatternFill, Alignment
import re
import openpyxl
from PIL import Image as PILImage
import io
from django.core.files.uploadedfile import InMemoryUploadedFile

# ── LOGIN ──────────────────────────────────────────
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'login.html')

# ── LOGOUT ─────────────────────────────────────────
def logout_view(request):
    logout(request)
    return redirect('login')

# ── DASHBOARD ──────────────────────────────────────
@login_required(login_url='login')
def dashboard(request):
    from django.db.models import Sum
    from django.utils import timezone
    from datetime import timedelta
    import re

    total_productos  = Producto.objects.count()
    stock_total      = Producto.objects.aggregate(total=Sum('stock'))['total'] or 0
    stock_bajo       = Producto.objects.filter(stock__lte=5).count()
    total_categorias = Categoria.objects.count()
    ultimos          = Producto.objects.order_by('-creado')[:5]
    alertas          = Producto.objects.filter(stock__lte=5).order_by('stock')

    # Stock por categoría para las barras
    categorias = Categoria.objects.all()
    stock_por_categoria = []
    for c in categorias:
        total = Producto.objects.filter(categoria=c).aggregate(
            total=Sum('stock'))['total'] or 0
        stock_por_categoria.append({'nombre': c.nombre, 'total': total})
    stock_maximo = max([i['total'] for i in stock_por_categoria], default=1)

    # Ventas por día — últimos 5 días
    hoy = localtime(timezone.now()).date()
    DIAS_ES = {
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
    }

    ventas_por_dia = []
    for i in range(0, 5):
        dia = hoy - timedelta(days=i)
        total_vendido = 0
        registros = Historial.objects.filter(accion='venta')
        for r in registros:
            fecha_local = localtime(r.fecha).date()
            if fecha_local == dia:
                match = re.search(r'Vendió (\d+) unidad', r.descripcion)
                if match:
                    total_vendido += int(match.group(1))
        ventas_por_dia.append({
            'dia':   'Hoy' if i == 0 else DIAS_ES.get(dia.strftime('%A'), dia.strftime('%A')),
            'total': total_vendido,
        })

    # Productos más vendidos — top 5
    from collections import defaultdict
    conteo = defaultdict(int)
    registros_venta = Historial.objects.filter(accion='venta')
    for r in registros_venta:
        match = re.search(r'Vendió (\d+) unidad', r.descripcion)
        if match:
            conteo[r.producto] += int(match.group(1))
    mas_vendidos = sorted(conteo.items(), key=lambda x: x[1], reverse=True)[:5]

    return render(request, 'dashboard.html', {
        'total_productos':     total_productos,
        'stock_total':         stock_total,
        'stock_bajo':          stock_bajo,
        'total_categorias':    total_categorias,
        'ultimos':             ultimos,
        'alertas':             alertas,
        'stock_por_categoria': stock_por_categoria,
        'stock_maximo':        stock_maximo,
        'ventas_por_dia':      ventas_por_dia,
        'mas_vendidos':        mas_vendidos,
    })

# ── PRODUCTOS ──────────────────────────────────────
@login_required(login_url='login')
def lista_productos(request):
    productos  = Producto.objects.select_related('categoria').all()
    categorias = Categoria.objects.all()

    q          = request.GET.get('q', '')
    categoria  = request.GET.get('categoria', '')
    talle      = request.GET.get('talle', '')
    stock_bajo = request.GET.get('stock_bajo', '')

    if q:
        productos = productos.filter(nombre__icontains=q)
    if categoria:
        productos = productos.filter(categoria__id=categoria)
    if talle:
        productos = productos.filter(talle=talle)
    if stock_bajo:
        productos = productos.filter(stock__lte=5)

    # Paginación de a 10
    paginator = Paginator(productos, 10)
    page      = request.GET.get('page', 1)
    productos = paginator.get_page(page)

    return render(request, 'lista.html', {
        'productos':   productos,
        'categorias':  categorias,
        'talles':      Producto.TALLES,
        'filtros': {
            'q':          q,
            'categoria':  categoria,
            'talle':      talle,
            'stock_bajo': stock_bajo,
        }
    })

@login_required(login_url='login')
def crear_producto(request):
    form = ProductoForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        producto = form.save(commit=False)
        if 'foto' in request.FILES:
            producto.foto = comprimir_imagen(request.FILES['foto'])
        producto.save()
        Historial.objects.create(
            usuario     = request.user,
            accion      = 'crear',
            producto    = producto.nombre,
            descripcion = f"Talle: {producto.talle} | Color: {producto.color} | Stock: {producto.stock}"
        )
        messages.success(request, '✅ Producto creado correctamente')
        return redirect('lista_productos')
    return render(request, 'formulario.html', {'form': form, 'titulo': '➕ Nuevo Producto'})

@login_required(login_url='login')
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    foto_anterior = producto.foto

    form = ProductoForm(request.POST or None, request.FILES or None, instance=producto)
    if form.is_valid():
        if 'foto' in request.FILES and foto_anterior:
            import os
            if os.path.isfile(foto_anterior.path):
                os.remove(foto_anterior.path)
            producto = form.save(commit=False)
            producto.foto = comprimir_imagen(request.FILES['foto'])
            producto.save()
        else:
            form.save()
        Historial.objects.create(
            usuario     = request.user,
            accion      = 'editar',
            producto    = producto.nombre,
            descripcion = f"Talle: {producto.talle} | Color: {producto.color} | Stock: {producto.stock}"
        )
        messages.success(request, '✅ Producto actualizado correctamente')
        return redirect('lista_productos')
    return render(request, 'formulario.html', {'form': form, 'titulo': '✏️ Editar Producto'})

@login_required(login_url='login')
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        Historial.objects.create(
            usuario     = request.user,
            accion      = 'eliminar',
            producto    = producto.nombre,
            descripcion = f"Talle: {producto.talle} | Color: {producto.color} | Stock: {producto.stock}"
        )
        producto.delete()
        messages.success(request, '🗑️ Producto eliminado correctamente')
        return redirect('lista_productos')
    return render(request, 'confirmar_eliminar.html', {'producto': producto})

@login_required(login_url='login')
def exportar_excel(request):
    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Estilo del encabezado
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="570df8", end_color="570df8", fill_type="solid")
    header_align = Alignment(horizontal="center")

    # Encabezados
    columnas = ['ID', 'Nombre', 'Categoría', 'Talle', 'Color', 'Precio', 'Stock', 'Fecha']
    for col_num, columna in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=columna)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # Ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 20

    # Datos
    productos = Producto.objects.select_related('categoria').all()
    for row_num, producto in enumerate(productos, 2):
        ws.cell(row=row_num, column=1, value=producto.id)
        ws.cell(row=row_num, column=2, value=producto.nombre)
        ws.cell(row=row_num, column=3, value=producto.categoria.nombre if producto.categoria else '')
        ws.cell(row=row_num, column=4, value=producto.talle)
        ws.cell(row=row_num, column=5, value=producto.color)
        ws.cell(row=row_num, column=6, value=float(producto.precio))
        ws.cell(row=row_num, column=7, value=producto.stock)
        ws.cell(row=row_num, column=8, value=producto.creado.strftime('%d/%m/%Y %H:%M'))

    # Devolver el archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='login')
def lista_usuarios(request):
    if not request.user.is_staff:
        return redirect('dashboard')
    usuarios = User.objects.all().order_by('username')
    return render(request, 'usuarios/lista_usuarios.html', {'usuarios': usuarios})

@login_required(login_url='login')
def crear_usuario(request):
    if not request.user.is_staff:
        return redirect('dashboard')
    form = UsuarioForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, '✅ Usuario creado correctamente')
        return redirect('lista_usuarios')
    return render(request, 'usuarios/formulario_usuario.html', {
        'form': form,
        'titulo': '➕ Nuevo Usuario'
    })

@login_required(login_url='login')
def eliminar_usuario(request, pk):
    if not request.user.is_staff:
        return redirect('dashboard')
    usuario = get_object_or_404(User, pk=pk)
    if usuario == request.user:
        messages.error(request, '❌ No podés eliminarte a vos mismo')
        return redirect('lista_usuarios')
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, '🗑️ Usuario eliminado correctamente')
        return redirect('lista_usuarios')
    return render(request, 'usuarios/confirmar_eliminar_usuario.html', {'usuario': usuario})

@login_required(login_url='login')
def historial(request):
    if not request.user.is_staff:
        return redirect('dashboard')
    registros = Historial.objects.select_related('usuario').all()[:50]
    return render(request, 'historial.html', {'registros': registros})

@login_required(login_url='login')
def registrar_venta(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0))
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0')
            return redirect('lista_productos')
        if cantidad > producto.stock:
            messages.error(request, f'❌ Solo hay {producto.stock} unidades disponibles')
            return redirect('lista_productos')
        producto.stock -= cantidad
        producto.save()
        Historial.objects.create(
            usuario     = request.user,
            accion      = 'venta',
            producto    = producto.nombre,
            descripcion = f"Vendió {cantidad} unidad(es) · Talle: {producto.talle} · Color: {producto.color} · Stock restante: {producto.stock}"
        )
        messages.success(request, f'✅ Venta registrada — {cantidad} unidad(es) de {producto.nombre}')
        return redirect('lista_productos')
    return render(request, 'registrar_venta.html', {'producto': producto})

def comprimir_imagen(imagen, max_size=(800, 800), quality=75):
    img = PILImage.open(imagen)
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')
    img.thumbnail(max_size, PILImage.LANCZOS)
    output = io.BytesIO()
    img.save(output, format='JPEG', quality=quality, optimize=True)
    output.seek(0)
    return InMemoryUploadedFile(
        output, 'ImageField',
        f"{imagen.name.split('.')[0]}.jpg",
        'image/jpeg',
        output.getbuffer().nbytes,
        None
    )